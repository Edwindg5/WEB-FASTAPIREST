"""Router admin — reportes PDF/Excel. Requiere rol=administrador."""
from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import FileResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, text
from typing import Dict, Any, Optional
from datetime import datetime
import os
import io

from app.infrastructure.db.database import get_db
from app.infrastructure.db.models.reporte import ReporteModel
from app.infrastructure.db.models.lote_cafe import LoteCafeModel
from app.infrastructure.db.models.usuario import UsuarioModel
from app.core.security import get_current_admin_user

router = APIRouter(prefix="/admin/reportes", tags=["Admin — Reportes"])

REPORTS_DIR = "reports"
os.makedirs(REPORTS_DIR, exist_ok=True)


def _generar_pdf(lote_data: dict, estadisticas: dict, alertas: list) -> bytes:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(f"Reporte de Lote: {lote_data.get('nombre_lote', 'N/A')}", styles["Title"]))
    story.append(Spacer(1, 0.25 * inch))

    info = [
        ["ID Lote", str(lote_data.get("id_lote", ""))],
        ["Estado", lote_data.get("estado", "")],
        ["Variedad", lote_data.get("variedad", "") or ""],
        ["Tipo Proceso", lote_data.get("tipo_proceso", "") or ""],
        ["Peso (kg)", str(lote_data.get("peso_kg", ""))],
        ["Ubicación", lote_data.get("ubicacion", "") or ""],
    ]
    t = Table(info, colWidths=[2.5 * inch, 4 * inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.lightblue),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.25 * inch))

    story.append(Paragraph("Estadísticas Ambientales", styles["Heading2"]))
    est = [
        ["Métrica", "Valor"],
        ["Temp. Promedio", f"{estadisticas.get('temp_avg', 0):.1f} °C"],
        ["Humedad Promedio", f"{estadisticas.get('hum_avg', 0):.1f} %"],
        ["Total Lecturas", str(estadisticas.get("total_lecturas", 0))],
    ]
    te = Table(est, colWidths=[3 * inch, 3 * inch])
    te.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    story.append(te)
    story.append(Spacer(1, 0.25 * inch))

    story.append(Paragraph("Alertas del Lote", styles["Heading2"]))
    if alertas:
        al_data = [["Tipo", "Severidad", "Atendida"]]
        for a in alertas[:20]:
            al_data.append([str(a.get("tipo_alerta", "")), str(a.get("nivel_severidad", "")), str(a.get("atendida", ""))])
        ta = Table(al_data, colWidths=[3 * inch, 2 * inch, 1.5 * inch])
        ta.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.red),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))
        story.append(ta)
    else:
        story.append(Paragraph("Sin alertas registradas.", styles["Normal"]))

    story.append(Spacer(1, 0.5 * inch))
    story.append(Paragraph("Recomendaciones: Mantener temperatura y humedad dentro de rangos óptimos.", styles["Normal"]))

    doc.build(story)
    buffer.seek(0)
    return buffer.read()


def _generar_excel(lote_data: dict, estadisticas: dict, alertas: list) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Lote"
    ws["A1"] = f"Reporte Lote: {lote_data.get('nombre_lote', 'N/A')}"
    ws["A1"].font = Font(bold=True, size=14)

    datos = [
        ("ID Lote", lote_data.get("id_lote", "")),
        ("Nombre", lote_data.get("nombre_lote", "")),
        ("Estado", lote_data.get("estado", "")),
        ("Variedad", lote_data.get("variedad", "")),
        ("Tipo Proceso", lote_data.get("tipo_proceso", "")),
        ("Peso (kg)", lote_data.get("peso_kg", "")),
        ("Ubicación", lote_data.get("ubicacion", "")),
    ]
    for i, (k, v) in enumerate(datos, start=3):
        ws[f"A{i}"] = k
        ws[f"B{i}"] = v

    row = len(datos) + 5
    ws[f"A{row}"] = "Estadísticas"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1
    for k, v in [("Temp. Prom", estadisticas.get("temp_avg", 0)), ("Hum. Prom", estadisticas.get("hum_avg", 0)), ("Lecturas", estadisticas.get("total_lecturas", 0))]:
        ws[f"A{row}"] = k
        ws[f"B{row}"] = v
        row += 1

    row += 1
    ws[f"A{row}"] = "Alertas"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1
    for col, h in enumerate(["Tipo", "Severidad", "Atendida"], 1):
        ws.cell(row=row, column=col, value=h).font = Font(bold=True)
    row += 1
    for a in alertas[:50]:
        ws.cell(row=row, column=1, value=a.get("tipo_alerta", ""))
        ws.cell(row=row, column=2, value=a.get("nivel_severidad", ""))
        ws.cell(row=row, column=3, value=str(a.get("atendida", "")))
        row += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


@router.get("", summary="Listar todos los reportes")
async def listar_reportes(
    formato: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    current_user: Dict[str, Any] = Depends(get_current_admin_user),
):
    query = select(ReporteModel)
    if formato:
        query = query.where(ReporteModel.formato == formato)

    total = (await db.execute(select(func.count()).select_from(query.subquery()))).scalar_one()
    offset = (page - 1) * limit
    result = await db.execute(query.order_by(ReporteModel.fecha_generacion.desc()).offset(offset).limit(limit))
    reportes = result.scalars().all()

    items = []
    for r in reportes:
        lote_r = await db.execute(select(LoteCafeModel.nombre_lote).where(LoteCafeModel.id_lote == r.id_lote))
        lote_nombre = lote_r.scalar_one_or_none()
        usr_r = await db.execute(select(UsuarioModel.nombre).where(UsuarioModel.id_usuario == r.id_usuario))
        usuario_nombre = usr_r.scalar_one_or_none()
        items.append({
            "id_reporte": r.id_reporte,
            "id_lote": r.id_lote,
            "lote_nombre": lote_nombre,
            "tipo_reporte": r.tipo_reporte,
            "formato": r.formato,
            "url_archivo": r.url_archivo,
            "usuario_nombre": usuario_nombre,
            "fecha_generacion": r.fecha_generacion,
        })
    return {"total": total, "page": page, "limit": limit, "items": items}


@router.post("/generar", status_code=status.HTTP_201_CREATED, summary="Generar reporte PDF o Excel")
async def generar_reporte(
    body: dict,
    db: AsyncSession = Depends(get_db),
    current_user: Dict[str, Any] = Depends(get_current_admin_user),
):
    id_lote = body.get("id_lote")
    tipo_reporte = body.get("tipo_reporte", "secado_completo")
    formato = body.get("formato", "pdf").lower()

    if formato not in ("pdf", "excel"):
        raise HTTPException(status_code=400, detail="Formato debe ser 'pdf' o 'excel'")

    lote_r = await db.execute(select(LoteCafeModel).where(LoteCafeModel.id_lote == id_lote))
    lote = lote_r.scalar_one_or_none()
    if not lote:
        raise HTTPException(status_code=404, detail="Lote no encontrado")

    lote_data = {
        "id_lote": lote.id_lote,
        "nombre_lote": lote.nombre_lote,
        "estado": lote.estado,
        "variedad": lote.variedad,
        "tipo_proceso": lote.tipo_proceso,
        "peso_kg": float(lote.peso_kg) if lote.peso_kg else None,
        "ubicacion": lote.ubicacion,
    }

    est_r = await db.execute(
        text("SELECT AVG(temperatura) ta, AVG(humedad) ha, COUNT(*) cnt "
             "FROM lecturas_ambientales la "
             "JOIN lotes_cafe l ON la.id_sensor = l.id_sensor "
             "WHERE l.id_lote = :lid"),
        {"lid": id_lote},
    )
    er = est_r.one()
    estadisticas = {"temp_avg": float(er.ta or 0), "hum_avg": float(er.ha or 0), "total_lecturas": int(er.cnt or 0)}

    alertas_r = await db.execute(
        text("SELECT tipo_alerta, nivel_severidad, atendida FROM alertas WHERE id_lote = :lid LIMIT 50"),
        {"lid": id_lote},
    )
    alertas = [{"tipo_alerta": r.tipo_alerta, "nivel_severidad": r.nivel_severidad, "atendida": r.atendida}
               for r in alertas_r.all()]

    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    ext = "pdf" if formato == "pdf" else "xlsx"
    filename = f"reporte_lote_{id_lote}_{tipo_reporte}_{ts}.{ext}"
    filepath = os.path.join(REPORTS_DIR, filename)

    content = _generar_pdf(lote_data, estadisticas, alertas) if formato == "pdf" else _generar_excel(lote_data, estadisticas, alertas)
    with open(filepath, "wb") as f:
        f.write(content)

    reporte = ReporteModel(
        id_lote=id_lote,
        id_usuario=int(current_user.get("sub")),
        tipo_reporte=tipo_reporte,
        formato=formato,
        url_archivo=filepath,
        fecha_generacion=datetime.utcnow(),
    )
    db.add(reporte)
    await db.commit()
    await db.refresh(reporte)

    return {
        "id_reporte": reporte.id_reporte,
        "id_lote": reporte.id_lote,
        "tipo_reporte": reporte.tipo_reporte,
        "formato": reporte.formato,
        "url_archivo": reporte.url_archivo,
        "fecha_generacion": reporte.fecha_generacion,
    }


@router.get("/{id}/descargar", summary="Descargar archivo de reporte")
async def descargar_reporte(
    id: int,
    db: AsyncSession = Depends(get_db),
    current_user: Dict[str, Any] = Depends(get_current_admin_user),
):
    result = await db.execute(select(ReporteModel).where(ReporteModel.id_reporte == id))
    reporte = result.scalar_one_or_none()
    if not reporte:
        raise HTTPException(status_code=404, detail="Reporte no encontrado")
    if not reporte.url_archivo or not os.path.exists(reporte.url_archivo):
        raise HTTPException(status_code=404, detail="Archivo no encontrado")

    media_type = "application/pdf" if reporte.formato == "pdf" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return FileResponse(path=reporte.url_archivo, media_type=media_type, filename=os.path.basename(reporte.url_archivo))
